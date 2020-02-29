# mypy: warn_unreachable = False

from typing import Dict, Tuple

import argparse
import json
import sys
import copy
import pathlib

import openpyxl  # type: ignore

from dp.stringify_csv import Csvify
from dp.load_course import load_course

DEFAULT_STYLE = 'default'
default = openpyxl.styles.NamedStyle(DEFAULT_STYLE)
default.font = openpyxl.styles.Font(name='Calibri', size=11, bold=False)
default.alignment = openpyxl.styles.Alignment(wrap_text=True)

done_thin = openpyxl.styles.Side(border_style="thin", color="006100")
pending_thin = openpyxl.styles.Side(border_style="thin", color="000000")
partial_thin = openpyxl.styles.Side(border_style="thin", color="9C675D")
failed_thin = openpyxl.styles.Side(border_style="thin", color="FFFFFF")

DONE_STYLE = 'done'
done = openpyxl.styles.NamedStyle(DONE_STYLE)
done.font = copy.copy(default.font)
done.font.color = "FF006100"
done.fill = openpyxl.styles.PatternFill(fill_type="solid", start_color='FFC6EFCE', end_color='FFC6EFCE')
done.border = openpyxl.styles.Border(top=done_thin, left=done_thin, right=done_thin, bottom=done_thin)
done.alignment = default.alignment

PENDING_STYLE = 'pending'
pending = openpyxl.styles.NamedStyle(PENDING_STYLE)
pending.font = copy.copy(default.font)
pending.fill = openpyxl.styles.PatternFill(fill_type="solid", start_color='FFB7DEE8', end_color='FFB7DEE8')
pending.border = openpyxl.styles.Border(top=pending_thin, left=pending_thin, right=pending_thin, bottom=pending_thin)
pending.alignment = default.alignment

PARTIAL_STYLE = 'partial'
partial = openpyxl.styles.NamedStyle(PARTIAL_STYLE)
partial.font = copy.copy(default.font)
partial.font.color = "FF9C675D"
partial.fill = openpyxl.styles.PatternFill(fill_type="solid", start_color='FFFFEB9C', end_color='FFFFEB9C')
partial.border = openpyxl.styles.Border(top=partial_thin, left=partial_thin, right=partial_thin, bottom=partial_thin)
partial.alignment = default.alignment

FAILED_STYLE = 'failed'
failed = openpyxl.styles.NamedStyle(FAILED_STYLE)
failed.font = copy.copy(default.font)
failed.font.color = "FFFFFFFF"
failed.font.bold = True
failed.fill = openpyxl.styles.PatternFill(fill_type="solid", start_color='FF000000', end_color='FF000000')
failed.border = openpyxl.styles.Border(top=partial_thin, left=partial_thin, right=partial_thin, bottom=partial_thin)
failed.alignment = default.alignment

HEADER_STYLE = 'header'
header = openpyxl.styles.NamedStyle(HEADER_STYLE)
header.font = copy.copy(default.font)
# header.font.bold = True
# header.font.size = 10
# header.font.name = 'JetBrains Mono'
header.alignment = openpyxl.styles.Alignment(vertical='top', wrap_text=True)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("out_file")
    parser.add_argument("json_data", nargs="*")
    args = parser.parse_args()

    wb = openpyxl.Workbook()
    wb.add_named_style(default)
    wb.add_named_style(header)
    wb.add_named_style(done)
    wb.add_named_style(pending)
    wb.add_named_style(partial)
    wb.add_named_style(failed)

    ws = wb.active
    ws.title = pathlib.Path(args.out_file).stem
    ws.freeze_panes = "B2"
    ws.sheet_view.zoomScale = 130

    rows = []
    for filename in args.json_data:
        with open(filename, 'r', encoding='utf-8') as infile:
            data = json.load(infile)

        transcript = [load_course(c) for c in data['transcript']]
        output = data['output']

        stnum = pathlib.Path(filename).stem.split('-')[0]
        csvifier = Csvify(stnum=stnum, transcript={c.clbid: c for c in transcript})

        print(stnum)

        rows.append(dict(csvifier.process(output)))

    dims: Dict[str, float] = {}
    for col, key in enumerate(rows[0].keys()):
        cell = ws.cell(value=key, column=col + 1, row=1)
        cell.style = HEADER_STYLE

        column = cell.column_letter
        if key == 'Status':
            dims[column] = 15
        else:
            longest_line = max(max(estimate_line_length(l) for l in key.split('\n')), 5)
            dims[column] = max(dims.get(column, 0), float(longest_line))

    for _col, width in dims.items():
        ws.column_dimensions[_col].width = width

    for row_n, row in enumerate(rows):
        # plus 2, because Excel is 1-indexed, and the header row is row 1
        row_n += 2

        for col_n, cell_value in enumerate(row.values()):
            # add 1 because Excel is 1-indexed
            col_n += 1

            if col_n <= 5:
                style = DEFAULT_STYLE
            else:
                style, cell_value = compute_style(cell_value)

            cell = ws.cell(value=cell_value, column=col_n, row=row_n)
            cell.style = style

    # red_fill = openpyxl.styles.PatternFill(bgColor="FFC7CE")
    # dxf = openpyxl.styles.DifferentialStyle(fill=red_fill)
    # r = openpyxl.styles.Rule(type="expression", dxf=dxf, stopIfTrue=True)
    # r.formula = ['$A2="Microsoft"']
    # ws.conditional_formatting.add("A:ZZ", r)

    wb.save(args.out_file)

    return 0


def estimate_line_length(s: str) -> float:
    max_len = len(s)

    alpha_count = sum(1 for c in s if c.isalpha())
    digit_count = sum(1 for c in s if c.isdigit())

    max_len -= alpha_count
    max_len -= digit_count

    estimate = digit_count + (max_len * 0.35) + 1

    if estimate > 100:
        estimate += alpha_count * 0.9
        x = len(s)
        print(f"{x} :: {estimate:,} :: {s}")
    else:
        estimate += alpha_count

    return estimate


def compute_style(value: str) -> Tuple[str, str]:
    if value.startswith('done:'):
        return DONE_STYLE, value

    if value.startswith('pending-current:'):
        value = value.replace('pending-current:', 'in-progress:')
        return PENDING_STYLE, value

    if value.startswith('needs-more-items:'):
        value = value.replace('needs-more-items:', 'partial:')
        return PARTIAL_STYLE, value

    if value.startswith('failed-invariant:'):
        return FAILED_STYLE, value

    if ':' not in value:
        if value:
            return DONE_STYLE, value
        else:
            return DEFAULT_STYLE, value

    return DEFAULT_STYLE, value


if __name__ == "__main__":
    sys.exit(main())
    # wb = openpyxl.load_workbook('./csci.xlsx')
    # ws = wb.active
    # for r in ws.conditional_formatting:
    #     print(r)
    #     print(dir(r))
    #     print(r.rules[0].formula)
